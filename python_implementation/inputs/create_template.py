"""Generate the inputs Excel template with dropdown validations.

Run once to create/reset the template:
    python -m python_implementation.inputs.create_template
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GROUP_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
GROUP_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
VALUE_FONT = Font(name="Calibri", size=11)
LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Input definitions: (param_name, display_name, default, value_type, options, description)
#   value_type: "dropdown", "int", "float", "bool"
#   options: list of strings for dropdown, None for numeric/bool
# ---------------------------------------------------------------------------

INPUTS = [
    # --- Data File ---
    ("__group__", "Data File", None, None, None, None),
    ("data_file_path", "Data File Path", "", "text", None,
     "Path to OHLCV data file (CSV or Excel)"),

    # --- Pattern Type ---
    ("__group__", "Pattern Type", None, None, None, None),
    ("pattern_type", "Pattern Type", "XABCD", "dropdown",
     ["XAB", "XABC", "XABCD", "XABCDE", "XABCDEF"],
     "Number of points in the pattern"),
    ("pattern_direction", "Pattern Direction", "Both", "dropdown",
     ["Bullish", "Bearish", "Both"],
     "Filter by pattern direction"),

    # --- Length Properties ---
    ("__group__", "Length Properties", None, None, None, None),
    ("b_min", "B Min", 20, "int", None,
     "Minimum bars between X and B"),
    ("b_max", "B Max", 100, "int", None,
     "Maximum bars between X and B"),
    ("max_search_bars", "Max Search Bars", 0, "int", None,
     "Maximum bars to search back (0 = all history)"),
    ("px_length_percentage", "PX Length %", 10.0, "float", None,
     "P-to-X length as percentage of X-to-B distance"),
    ("min_b_to_c_btw_x_b", "Min B-to-C % of XB", 0.0, "float", None,
     "Minimum B-to-C distance as % of XB bar count"),
    ("max_b_to_c_btw_x_b", "Max B-to-C % of XB", 100.0, "float", None,
     "Maximum B-to-C distance as % of XB bar count"),
    ("min_c_to_d_btw_x_b", "Min C-to-D % of XB", 0.0, "float", None,
     "Minimum C-to-D distance as % of XB bar count"),
    ("max_c_to_d_btw_x_b", "Max C-to-D % of XB", 100.0, "float", None,
     "Maximum C-to-D distance as % of XB bar count"),
    ("min_d_to_e_btw_x_b", "Min D-to-E % of XB", 0.0, "float", None,
     "Minimum D-to-E distance as % of XB bar count"),
    ("max_d_to_e_btw_x_b", "Max D-to-E % of XB", 100.0, "float", None,
     "Maximum D-to-E distance as % of XB bar count"),
    ("min_e_to_f_btw_x_b", "Min E-to-F % of XB", 0.0, "float", None,
     "Minimum E-to-F distance as % of XB bar count"),
    ("max_e_to_f_btw_x_b", "Max E-to-F % of XB", 100.0, "float", None,
     "Maximum E-to-F distance as % of XB bar count"),

    # --- Retracement Properties ---
    ("__group__", "Retracement Properties", None, None, None, None),
    ("max_width_percentage", "Max A Width %", 100.0, "float", None,
     "Maximum A deviation from XB line (%)"),
    ("min_width_percentage", "Min A Width %", 0.0, "float", None,
     "Minimum A deviation from XB line (%)"),
    ("x_to_a_b_max", "B Retracement Max %", 100.0, "float", None,
     "Maximum B retracement of XA range (%)"),
    ("x_to_a_b_min", "B Retracement Min %", -100.0, "float", None,
     "Minimum B retracement of XA range (%)"),

    # --- Dynamic Height ---
    ("__group__", "Dynamic Height", None, None, None, None),
    ("every_increasing_of_value", "Height Step (candles)", 5, "int", None,
     "For every N candles, height tolerance increases by 1 step"),
    ("width_increasing_percentage_x_to_b", "Height Increase % (XB)", 0.0, "float", None,
     "Percentage increase in XB height tolerance per step"),
    ("width_increasing_percentage_a_e", "Height Increase % (AE)", 0.0, "float", None,
     "Percentage increase in AE height tolerance per step"),

    # --- Validation ---
    ("__group__", "Validation", None, None, None, None),
    ("strict_xb_validation", "Strict XB Validation", "False", "dropdown",
     ["True", "False"],
     "Enable strict XB validation mode"),
    ("only_draw_most_recent", "Only Most Recent", "True", "dropdown",
     ["True", "False"],
     "Enforce minimum bar spacing between patterns"),
    ("min_bars_between_patterns", "Min Bars Between Patterns", 10, "int", None,
     "Minimum bars between last points of consecutive patterns"),
    ("slope_buffer_pct", "Slope Buffer %", 0.0, "float", None,
     "Buffer percentage for AC/BD/CE/DF slope validation (XB is always strict)"),

    # --- Channel Type ---
    ("__group__", "Channel Settings", None, None, None, None),
    ("channel_type", "Channel Type", "Parallel", "dropdown",
     ["Parallel", "Straight", "Non_Parallel", "All_Types"],
     "A-channel slope mode relative to XB slope"),
    ("xb_upper_width_pct", "XB Upper Width %", 0.5, "float", None,
     "XB channel upper band width (% of center price)"),
    ("xb_lower_width_pct", "XB Lower Width %", 0.5, "float", None,
     "XB channel lower band width (% of center price)"),
    ("a_upper_width_pct", "A Upper Width %", 0.5, "float", None,
     "A channel upper band width (% of center price)"),
    ("a_lower_width_pct", "A Lower Width %", 0.5, "float", None,
     "A channel lower band width (% of center price)"),
    ("channel_extension_bars", "Channel Extension Bars", 200, "int", None,
     "Bars past last point to extend channels visually"),

    # --- Golden Line ---
    ("__group__", "Golden Line Settings", None, None, None, None),
    ("f_percentage", "FG Separator Start %", 50.0, "float", None,
     "Starting FG separator height percentage"),
    ("fg_increasing_percentage", "FG Increment %", 5, "int", None,
     "FG separator increment percentage per iteration"),
    ("first_line_percentage", "FirstLine Slope %", 4.0, "float", None,
     "Initial FirstLine slope percentage"),
    ("first_line_decrease_percentage", "FirstLine Decrease %", 0.01, "float", None,
     "FirstLine slope decrement per iteration"),
    ("max_below_max_above_diff_percentage", "M/N Diff Tolerance %", 40.0, "float", None,
     "Maximum allowed difference between M and N deviations (%)"),
    ("mn_buffer_percent", "MN Buffer %", 0.0, "float", None,
     "Safety margin for MN golden line (% of M-N distance)"),
    ("mn_length_percent", "MN Min Length %", 0.0, "float", None,
     "Minimum MN segment length as % of SP-to-last distance (0=no min)"),
    ("mn_extension_bars", "MN Extension Bars", 20, "int", None,
     "Golden line extension bars beyond last point"),
    ("extension_break_close", "Extension Break on Close", "False", "dropdown",
     ["True", "False"],
     "Use CLOSE instead of HIGH/LOW for slope extension break detection"),

    # --- Dynamic Last Point ---
    ("__group__", "Dynamic Last Point", None, None, None, None),
    ("enable_dynamic_last_point", "Enable Dynamic Tracking", "True", "dropdown",
     ["True", "False"],
     "Enable dynamic last point tracking for live trading"),
    ("max_dynamic_iterations", "Max Dynamic Iterations", 10, "int", None,
     "Maximum number of dynamic point updates per pattern"),

    # --- Filters ---
    ("__group__", "Filters", None, None, None, None),
    ("divergence_type", "Divergence Type", "None", "dropdown",
     ["None", "Time", "Volume", "Time_Volume"],
     "Divergence filter mode"),
    ("tick_min_speed", "Tick Min Speed", 500000, "int", None,
     "Minimum tick speed threshold (seconds per bar)"),
]


def create_template(output_path: str | Path | None = None) -> Path:
    """Create the inputs Excel template with dropdown validations."""
    if output_path is None:
        output_path = Path(__file__).parent / "config.xlsx"
    output_path = Path(output_path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inputs"

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 55

    # Header row
    headers = ["Parameter", "Display Name", "Value", "Type", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    row = 2
    for entry in INPUTS:
        param, display, default, vtype, options, desc = entry

        if param == "__group__":
            # Group header row
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.fill = GROUP_FILL
                cell.border = THIN_BORDER
            ws.cell(row=row, column=1, value=display).font = GROUP_FONT
            ws.cell(row=row, column=2).font = GROUP_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.cell(row=row, column=1).alignment = LEFT
            row += 1
            continue

        # Parameter name (column A) - locked appearance
        cell_a = ws.cell(row=row, column=1, value=param)
        cell_a.font = VALUE_FONT
        cell_a.fill = LOCKED_FILL
        cell_a.border = THIN_BORDER
        cell_a.alignment = LEFT

        # Display name (column B) - locked appearance
        cell_b = ws.cell(row=row, column=2, value=display)
        cell_b.font = VALUE_FONT
        cell_b.fill = LOCKED_FILL
        cell_b.border = THIN_BORDER
        cell_b.alignment = LEFT

        # Value (column C) - editable
        cell_c = ws.cell(row=row, column=3, value=default)
        cell_c.font = Font(name="Calibri", size=11, bold=True)
        cell_c.border = THIN_BORDER
        cell_c.alignment = CENTER

        # Type (column D) - locked appearance
        cell_d = ws.cell(row=row, column=4, value=vtype)
        cell_d.font = VALUE_FONT
        cell_d.fill = LOCKED_FILL
        cell_d.border = THIN_BORDER
        cell_d.alignment = CENTER

        # Description (column E) - locked appearance
        cell_e = ws.cell(row=row, column=5, value=desc)
        cell_e.font = Font(name="Calibri", size=10, color="666666")
        cell_e.fill = LOCKED_FILL
        cell_e.border = THIN_BORDER
        cell_e.alignment = LEFT

        # Add dropdown validation for dropdown types
        if vtype == "dropdown" and options:
            formula = ",".join(options)
            dv = DataValidation(
                type="list",
                formula1=f'"{formula}"',
                allow_blank=False,
                showDropDown=False,
            )
            dv.prompt = desc
            dv.promptTitle = display
            dv.error = f"Please select from: {', '.join(options)}"
            dv.errorTitle = "Invalid Selection"
            cell_ref = f"C{row}"
            dv.add(cell_ref)
            ws.add_data_validation(dv)

        row += 1

    wb.save(output_path)
    print(f"Input template created: {output_path}")
    return output_path


if __name__ == "__main__":
    create_template()
